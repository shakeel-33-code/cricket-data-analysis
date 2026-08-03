#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler
players = pd.read_csv('c:/Users/shake/Downloads/data-analytics-project-for-beginners/t20_csv_files/t20_csv_files/dim_players.csv')
batting = pd.read_csv('c:/Users/shake/Downloads/data-analytics-project-for-beginners/t20_csv_files/t20_csv_files/fact_bating_summary.csv')
bowling = pd.read_csv('c:/Users/shake/Downloads/data-analytics-project-for-beginners/t20_csv_files/t20_csv_files/fact_bowling_summary.csv')
batting_stats = batting.groupby('batsmanName').agg({
    'runs': 'sum',
    'balls': 'sum',
    '4s': 'sum',
    '6s': 'sum',
    'match_id': 'nunique'  
}).reset_index()
batting_stats.rename(columns={'match_id': 'batting_matches'}, inplace=True)
batting_stats['batting_avg'] = batting_stats['runs'] / batting_stats['batting_matches']
batting_stats['strike_rate'] = (batting_stats['runs'] / batting_stats['balls']) * 100
bowling_stats = bowling.groupby('bowlerName').agg({
    'wickets': 'sum',
    'runs': 'sum',
    'overs': 'sum',
    'match_id': 'nunique'  
}).reset_index()
bowling_stats.rename(columns={'match_id': 'bowling_matches'}, inplace=True)
bowling_stats['economy'] = bowling_stats['runs'] / bowling_stats['overs']
bowling_stats['bowling_avg'] = np.where(
    bowling_stats['wickets'] > 0,
    bowling_stats['runs'] / bowling_stats['wickets'],
    999  
)
complete_stats = players.merge(batting_stats, left_on='name', right_on='batsmanName', how='left').merge(
    bowling_stats, left_on='name', right_on='bowlerName', how='left', suffixes=('_batting', '_bowling'))
complete_stats = complete_stats.fillna(0)
scaler = MinMaxScaler()
batting_features = ['batting_avg', 'strike_rate']
complete_stats[batting_features] = scaler.fit_transform(complete_stats[batting_features])
boundary_stats = complete_stats[['4s', '6s']].sum(axis=1).values.reshape(-1, 1)
normalized_boundaries = scaler.fit_transform(boundary_stats)
bowling_features = ['wickets', 'economy']
complete_stats[bowling_features] = scaler.fit_transform(complete_stats[bowling_features])
complete_stats['batting_score'] = (
    complete_stats['batting_avg'] * 0.3 +
    complete_stats['strike_rate'] * 0.3 +
    normalized_boundaries.flatten() * 0.4
)
complete_stats['bowling_score'] = (
    complete_stats['wickets'] * 0.4 +
    (1 - complete_stats['economy']) * 0.3 +
    (1 - (complete_stats['bowling_avg'] / complete_stats['bowling_avg'].max())) * 0.3
)
def select_players(df, role, n, sort_by='batting_score'):
    return df[df['playingRole'].str.contains(role, na=False, case=False)].nlargest(n, sort_by)
best_wicketkeeper = select_players(complete_stats, 'Wicketkeeper', 1, 'batting_score')
best_batsmen = select_players(complete_stats, 'Batter|Order', 4, 'batting_score')
best_allrounders = select_players(complete_stats, 'Allrounder', 2, 'batting_score')
best_bowlers = select_players(complete_stats, 'Bowler', 4, 'bowling_score')
best_eleven = pd.concat([best_wicketkeeper, best_batsmen, best_allrounders, best_bowlers])
def format_stats(row):
    matches = max(row['batting_matches'], row['bowling_matches'])
    return {
        'Name': row['name'],
        'Team': row['team'],
        'Role': row['playingRole'],
        'Matches': int(matches),
        'Runs': int(row['runs_batting']) if 'runs_batting' in row else 0,
        'Batting Average': round(row['runs_batting'] / row['batting_matches'], 2) if row['batting_matches'] > 0 else 0,
        'Strike Rate': round(row['strike_rate'], 2),
        'Fours': int(row['4s']),
        'Sixes': int(row['6s']),
        'Wickets': int(row['wickets']) if 'wickets' in row else 0,
        'Economy': round(row['economy'], 2) if 'economy' in row else 0
    }
final_team = pd.DataFrame([format_stats(row) for _, row in best_eleven.iterrows()])
column_order = ['Name', 'Team', 'Role', 'Matches', 'Runs', 'Batting Average', 
                'Strike Rate', 'Fours', 'Sixes', 'Wickets', 'Economy']
final_team = final_team[column_order]
final_team.to_csv('best_playing_eleven1.csv', index=False)
print("\nBest Playing 11:\n")
print("\nWicketkeeper:")
print(final_team[final_team['Role'].str.contains('Wicketkeeper')].to_string(index=False))
print("\nBatsmen:")
print(final_team[final_team['Role'].str.contains('Batter|Order')].to_string(index=False))
print("\nAll-rounders:")
print(final_team[final_team['Role'].str.contains('Allrounder')].to_string(index=False))
print("\nBowlers:")
print(final_team[final_team['Role'].str.contains('Bowler')].to_string(index=False))


# In[2]:


import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler
import os
players = pd.read_csv('c:/Users/shake/Downloads/data-analytics-project-for-beginners/t20_csv_files/t20_csv_files/dim_players.csv')
batting = pd.read_csv('c:/Users/shake/Downloads/data-analytics-project-for-beginners/t20_csv_files/t20_csv_files/fact_bating_summary.csv')
bowling = pd.read_csv('c:/Users/shake/Downloads/data-analytics-project-for-beginners/t20_csv_files/t20_csv_files/fact_bowling_summary.csv')
def format_stats(row):
    matches = max(row['batting_matches'], row['bowling_matches'])
    return {
        'Name': row['name'],
        'Team': row['team'],
        'Role': row['playingRole'],
        'Matches': int(matches),
        'Runs': int(row['runs_batting']) if 'runs_batting' in row else 0,
        'Batting Average': round(row['runs_batting'] / row['batting_matches'], 2) if row['batting_matches'] > 0 else 0,
        'Fours': int(row['4s']),
        'Sixes': int(row['6s'])
    }
final_team = pd.DataFrame([format_stats(row) for _, row in best_eleven.iterrows()])
column_order = ['Name', 'Team', 'Role', 'Matches', 'Runs', 'Batting Average', 'Fours', 'Sixes']
final_team = final_team[column_order]
output_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'best_playing_eleven.csv')
try:
    final_team.to_csv(output_path, index=False)
    print(f"\nFile successfully saved to: {output_path}")
except Exception as e:
    print(f"Error saving file: {e}")
    print("\nBest Playing 11 (DataFrame output):")
    print(final_team.to_string(index=False))
print("\nBest Playing 11:\n")
print("\nWicketkeeper:")
print(final_team[final_team['Role'].str.contains('Wicketkeeper')].to_string(index=False))
print("\nBatsmen:")
print(final_team[final_team['Role'].str.contains('Batter|Order')].to_string(index=False))
print("\nAll-rounders:")
print(final_team[final_team['Role'].str.contains('Allrounder')].to_string(index=False))
print("\nBowlers:")
print(final_team[final_team['Role'].str.contains('Bowler')].to_string(index=False))


# In[3]:


import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
output_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'best_playing_eleven1.xlsx')
final_team.to_excel(output_path, index=False, sheet_name='Best XI')
wb = load_workbook(output_path)
ws = wb.active
ws.insert_rows(1)
ws['A1'] = 'BEST PLAYING XI - T20 CRICKET TEAM'
ws.merge_cells('A1:H1')  # Merge cells across all columns for the headline
headline_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
data_font = Font(name='Arial', size=10)
headline_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_fill = PatternFill(start_color='757171', end_color='757171', fill_type='solid')
alt_row_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
headline_cell = ws['A1']
headline_cell.font = headline_font
headline_cell.fill = headline_fill
headline_cell.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, ws.max_column + 1):
    cell = ws.cell(2, col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
for row in range(3, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        if row % 2 == 0:
            cell.fill = alt_row_fill
for col in range(1, ws.max_column + 1):
    column_letter = get_column_letter(col)
    ws.column_dimensions[column_letter].width = 15
ws.row_dimensions[1].height = 30
ws.insert_rows(3)  # Insert a row for the first role heading
current_row = 3
roles = ['Wicketkeeper', 'Batsmen', 'All-rounders', 'Bowlers']
role_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
role_font = Font(name='Arial', size=11, bold=True)
for role in roles:
    ws.insert_rows(current_row)
    role_cell = ws.cell(current_row, 1)
    role_cell.value = role
    ws.merge_cells(f'A{current_row}:H{current_row}')
    role_cell.fill = role_fill
    role_cell.font = role_font
    role_cell.alignment = Alignment(horizontal='left', vertical='center')
    current_row += len(final_team[final_team['Role'].str.contains(role, case=False)]) + 1
wb.save(output_path)
print(f"\nFormatted Excel file has been created at: {output_path}")


# In[ ]:





# In[ ]:





# In[ ]:





# In[9]:





# In[10]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




